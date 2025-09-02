import streamlit as st
import os
import re
import warnings
import zipfile
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from copy import copy

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Define column mappings
mappings = {
    'B:E': 'B:E',
    'F:I': 'G:J',
    'J:R': 'L:T',
    'S': 'AC',
    'T:W': 'AE:AH',
    'AD': 'AK'
}

source_order = ['B:E', 'F:I', 'J:R', 'S', 'T:W', 'AD']

# Unique ID columns
UNIQUE_COL_SRC = 'AE'  # Employee ID source
UNIQUE_COL_DEST = 'AE'  # Employee ID destination

def get_columns_by_range(df, excel_range):
    if ':' in excel_range:
        start_col, end_col = excel_range.split(':')
        return df.iloc[:, column_index_from_string(start_col)-1:column_index_from_string(end_col)]
    else:
        return df.iloc[:, [column_index_from_string(excel_range)-1]]

def clean_string(s):
    if pd.isna(s):
        return s
    return re.sub(r'[^A-Za-z0-9]', '', str(s))

def extract_mapped_values(row):
    values = []
    for src_range in source_order:
        data = get_columns_by_range(pd.DataFrame([row]), src_range).values[0].tolist()
        if src_range == 'F:I':
            data[0] = clean_string(data[0])
        values.extend(data)
    return tuple(values)

def process_excel(input_file):
    # Load employee data
    employee_df = pd.read_excel(input_file, header=0)

    # Create in-memory ZIP file
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'a', zipfile.ZIP_DEFLATED, False) as zip_file:
        total_added = 0
        file_counts = {}

        chunk_size = 100
        existing_ids = set()
        file_index = 1
        row_count = 0

        # Create the first workbook
        wb = load_workbook('template.xlsx')
        ws = wb.active
        start_row = 10
        row_num = start_row

        for _, row in employee_df.iterrows():
            emp_id = str(row.iloc[column_index_from_string(UNIQUE_COL_SRC)-1]).strip()
            if emp_id in existing_ids:  # Skip duplicates completely
                continue

            existing_ids.add(emp_id)
            values = extract_mapped_values(row)
            val_index = 0

            for src_range, dest_range in mappings.items():
                if ':' in dest_range:
                    start_col, end_col = dest_range.split(':')
                    dest_cols = range(column_index_from_string(start_col), column_index_from_string(end_col)+1)
                else:
                    dest_cols = [column_index_from_string(dest_range)]

                for col_idx in dest_cols:
                    existing_cell = ws.cell(row=row_num, column=col_idx)
                    new_cell = ws.cell(row=row_num, column=col_idx, value=values[val_index])
                    if existing_cell.has_style:
                        new_cell.font = copy(existing_cell.font)
                        new_cell.border = copy(existing_cell.border)
                        new_cell.fill = copy(existing_cell.fill)
                        new_cell.number_format = copy(existing_cell.number_format)
                        new_cell.protection = copy(existing_cell.protection)
                        new_cell.alignment = copy(existing_cell.alignment)
                    
                    val_index += 1

            row_num += 1
            row_count += 1
            total_added += 1

            # If we reached 100, save file and reset
            if row_count == chunk_size:
                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_filename = f"output_{file_index}.xlsx"
                zip_file.writestr(excel_filename, excel_buffer.getvalue())
                file_counts[excel_filename] = row_count

                # Reset counters for next file
                file_index += 1
                row_count = 0
                wb = load_workbook('template.xlsx')
                ws = wb.active
                row_num = start_row

        # Save last file if it has rows
        if row_count > 0:
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_filename = f"output_{file_index}.xlsx"
            zip_file.writestr(excel_filename, excel_buffer.getvalue())
            file_counts[excel_filename] = row_count

    zip_buffer.seek(0)
    return zip_buffer, total_added, file_counts

def main():
    st.set_page_config(
    page_title="DVMSC",
    page_icon="logo.png",           
    layout="centered"
    )
    st.title("Security Bank Processor")
    st.write("Upload your file to process it according to the template.")
    st.write("The processed file will be available for download as a ZIP file containing of 100 limit data/rows per file.")
    st.write("Make sure your file have concatenated first name and last name and add it in AE column.")
    st.write("Formula for AE column: =CONCAT(B2,C2)")
    st.write("Make sure also to have 1 tab only")

    uploaded_file = st.file_uploader("Choose an Excel file", type=['xlsx', 'xls'])
    output_filename = st.text_input("Output ZIP filename (without .zip)", "output")

    if uploaded_file is not None and st.button("Process File"):
        try:
            with st.spinner('Processing your file...'):
                zip_buffer, total_added, branch_counts = process_excel(uploaded_file)
            
            st.success("File processed successfully!")
            st.download_button(
                label="Download ZIP file",
                data=zip_buffer,
                file_name=f"{output_filename}.zip",
                mime="application/zip"
            )
            
            st.write("### Processing Summary")
            st.write(f"Total batches processed: {len(branch_counts)}")
            st.write(f"Total rows added: {total_added}")
            for branch, count in branch_counts.items():
                st.write(f"- {branch}: {count} rows")
                
        except Exception as e:
            # st.error(f"Please check your column AE, and you have a formula for it.")
            st.error(f"An error occurred: {str(e)}")

if __name__ == '__main__':
    main()